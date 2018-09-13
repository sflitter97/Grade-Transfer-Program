import xlrd
import openpyxl


#VARIABLES THAT NEED TO BE CHANGED
#==============================================================================
importFile = ("import_file.xlsx")
inputGradeColumn = 3
exportWorkbook = openpyxl.load_workbook('export_file.xlsx')
outputGradeColumn = 7
#==============================================================================


#Import information
importWorkbook = xlrd.open_workbook(importFile)
importSheet = importWorkbook.sheet_by_index(0)

# Export information
sheets = exportWorkbook.sheetnames
exportSheet = exportWorkbook[sheets[0]]

# Output text files
unfStudFilename = ("unmatched_Students") + '.txt'
unfStudentsText = open(unfStudFilename, 'wb')
doubleEntriesFile = ("double_Entries") + '.txt'
dblEntrText = open(doubleEntriesFile, 'wb')


# This function creates a text file containing the names of students
def doubleSubmissions(name):
    dblEntrText.write(name.encode('utf-8') + b'\n')

# This function adds the names of students that were not found in the gradebook excel sheet to a list
def studentNotFound(name):    
    unfStudentsText.write(name.encode('utf-8') + b'\n')


# This function looks to see if a student from Stepik is in the canvas gradebook
def lookUpStudent(name, grade):
    for i in range(exportSheet.max_row):
        if exportSheet.cell(row=i + 1, column = 1).value == name:
            if exportSheet.cell(row=i + 1, column = outputGradeColumn).value:
                doubleSubmissions(name)
                return
            exportSheet.cell(row=i + 1, column = outputGradeColumn).value = grade
            return
    studentNotFound(name)


# "Main" part of program
for i in range(importSheet.nrows):
    name = importSheet.cell_value(i, 1) + ', ' + importSheet.cell_value(i, 2)
    grade = importSheet.cell_value(i, inputGradeColumn)
    lookUpStudent(name, grade)
    

# Closes and saves opened files
unfStudentsText.close()
dblEntrText.close()
exportWorkbook.save('export_file.xlsx')